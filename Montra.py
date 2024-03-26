import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter import *
from threading import Thread
import requests
from requests.auth import HTTPBasicAuth
from requests.exceptions import ConnectionError
import os
from datetime import datetime
import base64
import xml.etree.ElementTree as ET
import uuid
import time
import openpyxl
from cryptography.fernet import Fernet
from pathlib import Path
import configparser
import customtkinter
from PIL import Image, ImageTk

clave_cifrado= b'5eWYhZWF9OBQqiI6k2urPzWBAdj0WZ5lz-m-xGn2mJ4=' #Esta es la clave que utiliza el archivo .ini para encriptarlo y desencriptarlo
fernet = Fernet(clave_cifrado)

class ProcesadorArchivos:
    def __init__(self, root):
        self.root = root
        self.root.title("MONTRA")
        self.contraseña_verificada = False    #Estado de contraseña verificada o no
        self.configuracion_bloqueada = False  #Estado de bloqueo de configuración
        self.direcciones_mac_permitidas = ["B4-D5-BD-E8-AA-1C", "D8-EB-97-B8-7A-BB", "4C-44-5B-95-52-82","50-EB-71-D4-BB-70","52-EB-71-D4-BB-6F","50-EB-71-D4-BB-73", "BC-F1-71-F3-5F-60", "30-05-05-B8-BB-35", "30-05-05-B8-B4-69", "AC-1A-3D-11-17-2E", "6C-24-08-CB-BC-94", "BC-F1-71-F3-5F-5D", "BE-F1-71-F3-5F-5C", "BC-F1-71-F3-5F-5C"]  # Lista de direcciones MAC permitidas  # Reemplaza con la MAC permitida
        
        # Contadores para envío exitoso y fallido
        self.envio_exitoso = 0
        self.envio_fallido = 0
        #Creación de la ventana
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True)
        self.medicion_tab = ttk.Frame(self.notebook)
        self.configuracion_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.medicion_tab, text="WebService", state="normal")
        self.notebook.add(self.configuracion_tab, text="Configuración", state="disabled")
        self.root.protocol("WM_DELETE_WINDOW", self.cerrar_aplicacion) #Metodo que cierra todos los procesos al cerrar la app.

        self.imagenes() #Metodo para cargar las imagenes
        self.create_medicion_tab() #Metodo para cargar la pestaña de medición
        self.create_configuracion_tab() #Metodo para cargar la pestaña de configuración
        self.cargar_configuracion() #Metodo para cargar el archivo .ini

        self.ejecutar = True   # Variable para controlar la ejecución del programa
        self.error=False  # Variable para controlar si hay algun error en el programa
        
        self.root.after(10, self.cargar_icono)  # Programar la carga del icono después de 10 milisegundos
    
#Metodo para cargar el icono para que al ejecutar el programa no se vea un salto de ventanas
    def cargar_icono(self):
        self.root.iconbitmap("Icons/logo-montra.ico")

#Metodo para cerrar la aplicación correctamente.
    def cerrar_aplicacion(self):
        # Detener el hilo de procesamiento
        self.ejecutar = False
        self.exportar_log()
        # Cerrar la interfaz gráfica
        self.root.destroy()

#VERIFICACIÓN DE MAC
    # Metodo para Obtener mac_address
    def get_mac_address(self):
        mac = uuid.UUID(int=uuid.getnode()).hex[-12:]
        formatted_mac = ':'.join([mac[i:i+2] for i in range(0, 12, 2)])
        formatted_mac=formatted_mac.upper()
        formatted_mac = formatted_mac.replace(":", "-")
        return formatted_mac

#METODOS DE CONFIGURACIÓN .INI
    #Metodo para cargar la configuración .ini
    def cargar_configuracion(self):
        mac_actual = self.get_mac_address()  # Usa el método de obtener_mac() definido
        if mac_actual in self.direcciones_mac_permitidas:
            config = configparser.ConfigParser()
            config.read('config.ini')
            if 'Configuracion' in config:
                self.token_url.set(self.desencriptar(config['Configuracion'].get('Token_url', '')))
                self.api_url.set(self.desencriptar(config['Configuracion'].get('api_url', '')))
                self.url_api_image.set(self.desencriptar(config['Configuracion'].get('url_api_image', '')))
                self.client_id.set(self.desencriptar(config['Configuracion'].get('client_id', '')))
                self.client_secret.set(self.desencriptar(config['Configuracion'].get('client_secret', '')))
                self.username.set(self.desencriptar(config['Configuracion'].get('username', '')))
                self.password.set(self.desencriptar(config['Configuracion'].get('password', '')))
                self.carpeta_archivos.set(self.desencriptar(config['Configuracion'].get('carpeta_archivos', '')))
                self.carpeta_imagenes.set(self.desencriptar(config['Configuracion'].get('carpeta_imagenes', '')))
                self.contraseña = (self.desencriptar(config['Configuracion'].get('contraseña_adicional')) )# Obtener la contraseña
        else:
            mensaje = "Este software solo puede ejecutarse en una computadora autorizada."
            messagebox.showerror("Error", mensaje)
            self.cerrar_aplicacion()
            #root.destroy()  # Cierra la aplicación
    
    #Metodo para guardar los datos de la configuración .ini
    def guardar_configuracion(self):
        config = configparser.ConfigParser()
        config.read('config.ini')
        config['Configuracion']['Token_url'] = self.encriptar(self.token_url.get())
        config['Configuracion']['api_url'] = self.encriptar(self.api_url.get())
        config['Configuracion']['url_api_image'] = self.encriptar(self.url_api_image.get())
        config['Configuracion']['client_id'] = self.encriptar(self.client_id.get())
        config['Configuracion']['client_secret'] = self.encriptar(self.client_secret.get())
        config['Configuracion']['username'] = self.encriptar(self.username.get())
        config['Configuracion']['password'] = self.encriptar(self.password.get())
        config['Configuracion']['carpeta_archivos'] = self.encriptar(self.carpeta_archivos.get())
        config['Configuracion']['carpeta_imagenes'] = self.encriptar(self.carpeta_imagenes.get())
        config['Configuracion']['contraseña_adicional'] = self.encriptar((self.contraseña))

        with open('config.ini', 'w') as configfile:
            config.write(configfile)
    
    #Metodo que encripta los datos con la clave de cifrado
    def encriptar(self, valor):
        return fernet.encrypt(valor.encode()).decode()
    
    #Metodo que desencripta datos con la clave de cifrado
    def desencriptar(self, valor_cifrado):
        token = valor_cifrado.encode()
        return fernet.decrypt(token).decode()

    #Metodo para cargue de imagenes
    def imagenes(self):
        self.logo_montra = tk.PhotoImage(file="Icons/imagen_1.png")
        self.logo_montra = self.logo_montra.subsample(1, 1)
        self.logo_cubiscan = tk.PhotoImage(file="Icons/Cubiscan_logo.png")
        self.logo_cubiscan = self.logo_cubiscan.subsample(1, 1)
        
        self.logo_mavesa = tk.PhotoImage(file="Icons/imagen_4.png")
        self.logo_mavesa = self.logo_mavesa.subsample(2, 2)

#METODOS PARA INGRESO POR CONTRASEÑA
    #Pestaña de ingreso de contraseña
    def abrir_pestana_configuraciones(self):
        if not self.contraseña_verificada:
            self.notebook.select(self.configuracion_tab)
            self.dialogo_contraseña = tk.Toplevel()
            self.dialogo_contraseña.title("Acceso")
            self.dialogo_contraseña.geometry("300x120")
            self.dialogo_contraseña.resizable(False, False)
            
            etiqueta_contraseña = ttk.Label(self.dialogo_contraseña, text="Ingrese la contraseña:")
            etiqueta_contraseña.pack(pady=5)

            self.entrada_contraseña = ttk.Entry(self.dialogo_contraseña, show="*")
            self.entrada_contraseña.pack(pady=5)
            self.entrada_contraseña.focus_set()
            self.entrada_contraseña.bind("<Return>", lambda event: self.verificar_contraseña())

            boton_verificar = ttk.Button(self.dialogo_contraseña, text="Verificar", command=self.verificar_contraseña)
            boton_verificar.pack(pady=5)
            self.dialogo_contraseña.grab_set()

            # Configurar una acción cuando se cierra la ventana de verificación
            self.dialogo_contraseña.protocol("WM_DELETE_WINDOW", self.restablecer_campos_configuracion)
            
            # Bloquear campos de configuración hasta que se verifique la contraseña
            for child in self.configuracion_tab.winfo_children():
                if isinstance(child, ttk.Entry):
                    child.config(state="disabled")
            self.dialogo_contraseña.iconbitmap("Icons/logo-montra.ico")

    #Metodo para que la pestaña de configuración permanezca inactiva si no se puso la clave
    def restablecer_campos_configuracion(self):
        # Restablecer solo si se cierra la ventana emergente de verificación con la "X"
        if self.dialogo_contraseña:
            self.dialogo_contraseña.destroy()
            if self.notebook.index("current") == 1:  # Verifica si la pestaña actual es la de configuración
                self.notebook.select(0)  # Cambia a la pestaña de medición
                self.notebook.tab(1, state="disabled")  # Deshabilita la pestaña de configuración
                # Limpiar y desbloquear los campos de configuración
                self.entrada_contraseña.delete(0, tk.END)
                for child in self.configuracion_tab.winfo_children():
                    if isinstance(child, ttk.Entry):
                        child.config(state="normal")
                self.configuracion_bloqueada = True  # Bloquear la configuración nuevamente
            else:
                self.configuracion_bloqueada = False  # Si se cierra la ventana de configuración desde la pestaña de medición

    #Meotdo para verificar si la contraseña ingresada es la almacenada
    def verificar_contraseña(self):
        contraseña_ingresada = self.entrada_contraseña.get()
        if contraseña_ingresada == self.contraseña:
            self.notebook.tab(1, state="normal")  # Habilita la pestaña de configuración
            self.dialogo_contraseña.destroy()
            # Habilitar los campos de configuración
            for child in self.configuracion_tab.winfo_children():
                if isinstance(child, ttk.Entry):
                    child.config(state="normal")
            self.contraseña_verificada = True
        else:
            self.entrada_contraseña.delete(0, tk.END)
            self.dialogo_contraseña.destroy() 
            messagebox.showerror("Error", "Contraseña incorrecta")

    #Metodo para cambio de contraseña
    def abrir_ventana_cambio_contraseña(self):
        # Ventana emergente para cambiar la contraseña
        cambio_contraseña_window = tk.Toplevel(self.root)
        cambio_contraseña_window.title("Cambiar Contraseña")

        ttk.Label(cambio_contraseña_window, text="Contraseña Actual:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        contraseña_actual_entry = ttk.Entry(cambio_contraseña_window, show="*")
        contraseña_actual_entry.grid(row=0, column=1, padx=10, pady=5)

        ttk.Label(cambio_contraseña_window, text="Nueva Contraseña:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        nueva_contraseña_entry = ttk.Entry(cambio_contraseña_window, show="*")
        nueva_contraseña_entry.grid(row=1, column=1, padx=10, pady=5)

        ttk.Button(cambio_contraseña_window, text="Guardar", command=lambda: self.guardar_nueva_contraseña(contraseña_actual_entry.get(), nueva_contraseña_entry.get(), cambio_contraseña_window)).grid(row=2, columnspan=2, padx=10, pady=5)
        
        cambio_contraseña_window.bind("<Return>", lambda event: self.guardar_nueva_contraseña(contraseña_actual_entry.get(), nueva_contraseña_entry.get(), cambio_contraseña_window))
        
        cambio_contraseña_window.iconbitmap("Icons/logo-montra.ico")

    #Metodo para guardar la nueva contraseña
    def guardar_nueva_contraseña(self, contraseña_actual, nueva_contraseña, window):
        if contraseña_actual != self.contraseña:
            messagebox.showerror("Error", "La contraseña actual es incorrecta.")
            window.destroy()
        else:
            if nueva_contraseña:
                self.contraseña = nueva_contraseña
                messagebox.showinfo("Contraseña Cambiada", "La contraseña ha sido cambiada con éxito.")
                window.destroy()
                self.guardar_configuracion()  # Guardar la nueva contraseña en el archivo config.ini

#Creación de ventanas de interfaz
    #Creación de ventana principal
    def create_medicion_tab(self):
        # Agregar la pestaña de medición al notebook
        self.notebook.add(self.medicion_tab, text="WebService", state="normal")

        # Insertarla en una etiqueta.
        self.colorbackground= "lightgrey"
        self.background = ttk.Label(self.medicion_tab, background=self.colorbackground)
        self.background.grid(row=0, column=0, columnspan=21, rowspan=2, pady=(0,0), sticky="snew")

        self.label_imagen1 = ttk.Label(self.medicion_tab, image=self.logo_montra, background=self.colorbackground)
        self.label_imagen1.grid(row=0, column=0, columnspan=2, padx=(30,0), pady=(10,0))

        self.label_imagen3 = ttk.Label(self.medicion_tab, image=self.logo_cubiscan, background=self.colorbackground)
        self.label_imagen3.grid(row=1, column=0, columnspan=2, padx=(30,0), pady=(0,10))

        self.label_imagen2 = ttk.Label(self.medicion_tab, image=self.logo_mavesa, background=self.colorbackground)
        self.label_imagen2.grid(row=0, column=19, rowspan=2, columnspan=2, padx=(0,0), sticky="w")
        

        self.boton_iniciar = customtkinter.CTkButton(self.medicion_tab, text="Iniciar", border_color="#AFACAC", border_width=1,   corner_radius=1,font=("Helvetica", 16), text_color="#000000", fg_color="#FFFFFF", hover_color="#D9F3FF", text_color_disabled="#000000", width=120, height=45, compound="left", command=self.iniciar_proceso)
        self.boton_iniciar.grid(row=0, column=2, columnspan=3, padx=(75,55),  pady=5, stick="w")
        
        self.boton_detener = customtkinter.CTkButton(self.medicion_tab, text="Detener", border_color="#AFACAC", border_width=1,   corner_radius=1,font=("Helvetica", 16), text_color="#000000", fg_color="#FFFFFF", hover_color="#D9F3FF", text_color_disabled="#000000",width=120, height=45, compound="left", command=self.detener_proceso)
        self.boton_detener.grid(row=1, column=2, columnspan=2, sticky="e", padx=(75,55), pady=(0,10))


        # Botón "Configuraciones" 
        configuraciones_image = Image.open("Icons/configuraciones.png")
        configuraciones_image = configuraciones_image.resize((20, 20))
        configuraciones_icon = ImageTk.PhotoImage(configuraciones_image)
        boton_configuraciones = ttk.Button(self.medicion_tab, image=configuraciones_icon, command=self.abrir_pestana_configuraciones)
        boton_configuraciones.image = configuraciones_icon
        boton_configuraciones.grid(row=8, column=20, sticky="ne")

        # Crear la tabla para mostrar los datos
        columns = ('SKU', 'PackType' ,'UOM' , 'Cajas', 'Inner', 'Largo', 'Ancho', 'Alto', 'Peso', 'Fecha')
        self.tree = ttk.Treeview(self.medicion_tab, columns=columns, show='headings')

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column('SKU', width=150)
            self.tree.column('PackType', width=80)
            self.tree.column('UOM', width=50)
            self.tree.column('Cajas', width=40)
            self.tree.column('Inner', width=40)
            self.tree.column('Largo', width=50)
            self.tree.column('Ancho', width=50)
            self.tree.column('Alto', width=50)
            self.tree.column('Peso', width=50)
            self.tree.column('Fecha', width=130)

        self.tree.grid(row=3, column=0, columnspan=20, pady=(10,5), padx=(10,10))
        
        # Aplicar un estilo con bordes a la tabla
        style = ttk.Style()
        style.configure("Treeview", font=('Helvetica', 9), rowheight=20)
        style.configure("Treeview.Heading", font=('Helvetica', 9))
        style.configure("Treeview.Treeview", borderwidth=1)  # Esto añade bordes alrededor de cada celda
        
        # Crear barras de desplazamiento
        y_scroll = ttk.Scrollbar(self.medicion_tab, orient="vertical", command=self.tree.yview)
        y_scroll.grid(row=3, column=20, sticky='wns')
        self.tree.configure(yscrollcommand=y_scroll.set)

        ttk.Label(self.medicion_tab, text="Respuesta WebService:").grid(row=6, column=0, columnspan=2, padx=10, sticky="w")
        self.response_entry = tk.Text(self.medicion_tab, state="disabled", background="#FCFFD0", font=("Arial", 10))
        self.response_entry.config(width=20, height=5)
        self.response_entry.grid(row=7, column=0, columnspan=20, pady=5, padx=(10,10), sticky="nsew")

        self.label_envio_exitoso = ttk.Label(self.medicion_tab, text="Envío exitosos: 0", foreground="green")
        self.label_envio_exitoso.grid(row=8, column=0, columnspan=2, padx=(120,0), pady=(0,10))

        self.label_envio_fallido = ttk.Label(self.medicion_tab, text="Envíos fallidos: 0", foreground="red")
        self.label_envio_fallido.grid(row=8, column=19, rowspan=2, columnspan=2, padx=(0,0), sticky="w")

    #Creación de ventana de configuración
    def create_configuracion_tab(self):
        
        espaciado_lateral=(30,0)
        
        self.token_url = tk.StringVar()
        # Información de autenticación
        self.api_url = tk.StringVar()
        self.url_api_image = tk.StringVar()
        
        self.client_id = tk.StringVar()
        self.client_secret = tk.StringVar()
        self.username = tk.StringVar()
        self.password = tk.StringVar()

        # Ruta de la carpeta donde se encuentran los archivos txt
        self.carpeta_archivos = tk.StringVar()
        self.carpeta_imagenes= tk.StringVar()
        
        # Ruta de la carpeta "procesados"
        
        self.carpeta_procesados_data = "Procesados/Data"
        self.carpeta_procesados_data_e = "Procesados/Data/Errores/"
        self.carpeta_procesados_img = "Procesados/Images"
        self.carpeta_procesados_img_e = "Procesados/Images/Errores/"

        ttk.Label(self.configuracion_tab, text="AUTENTICACIÓN WEBSERVICE", font=("Helvetica", 13)).grid(row=0, column=0, columnspan=3, pady=(20, 5), sticky="w", padx= espaciado_lateral)
        
        ttk.Label(self.configuracion_tab, text="URL del Web Service:").grid(row=1, column=1, pady=5, sticky="w", padx= espaciado_lateral)
        url_entry = ttk.Entry(self.configuracion_tab, textvariable=self.api_url,  width=45)
        url_entry.grid(row=1, column=2, pady=5, sticky="w")
        
        ttk.Label(self.configuracion_tab, text="URL Imagen:").grid(row=2, column=1, pady=5, sticky="w", padx= espaciado_lateral)
        url_entry = ttk.Entry(self.configuracion_tab, textvariable=self.url_api_image,  width=45)
        url_entry.grid(row=2, column=2, pady=5, sticky="w")
        
        ttk.Label(self.configuracion_tab, text="Client ID:").grid(row=3, column=1, pady=5, sticky="w", padx= espaciado_lateral)
        client_id_entry = ttk.Entry(self.configuracion_tab, textvariable=self.client_id, width=45)
        client_id_entry.grid(row=3, column=2, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="Client Secret:").grid(row=4, column=1, pady=5, sticky="w", padx= espaciado_lateral)
        client_secret_entry = ttk.Entry(self.configuracion_tab, textvariable=self.client_secret, width=45)
        client_secret_entry.grid(row=4, column=2, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="Usuario:").grid(row=5, column=1, pady=5, sticky="w", padx= espaciado_lateral)
        username_entry = ttk.Entry(self.configuracion_tab, textvariable=self.username, width=45)
        username_entry.grid(row=5, column=2, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="Contraseña:").grid(row=6, column=1, pady=5, sticky="w", padx= espaciado_lateral)
        password_entry = ttk.Entry(self.configuracion_tab, textvariable=self.password, width=45)
        password_entry.grid(row=6, column=2, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="URL del token:").grid(row=7, column=1, pady=5, sticky="w", padx= espaciado_lateral)
        token_url_entry = ttk.Entry(self.configuracion_tab, textvariable=self.token_url, width=45)
        token_url_entry.grid(row=7, column=2, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="PROCESAMIENTO DE DATOS", font=("Helvetica", 13)).grid(row=8, column=1, columnspan=3, pady=(20, 5), sticky="w", padx= espaciado_lateral)
        ttk.Label(self.configuracion_tab, text="Carpeta Origen Data:").grid(row=9, column=1, pady=5, sticky="w", padx= espaciado_lateral)

        # Mostrar la imagen "folder.png" al lado del campo "Carpeta Origen"
        folder_image = Image.open("Icons/folder.png")
        folder_image = folder_image.resize((20, 20))  # Redimensiona la imagen
        folder_icon = ImageTk.PhotoImage(folder_image)
        
        # Asignar la función seleccionar_carpeta al evento de clic del botón
        def seleccionar_carpeta():
            selected_folder = filedialog.askdirectory()
            if selected_folder:
                carpeta_origen_entry.delete(0, tk.END)  # Borrar el contenido actual del Entry
                carpeta_origen_entry.insert(0, selected_folder)  # Insertar la nueva ruta seleccionada

        folder_label = ttk.Button(self.configuracion_tab, image=folder_icon, command=seleccionar_carpeta)
        folder_label.image = folder_icon
        folder_label.grid(row=9, column=4, padx=(5, 0), pady=5, sticky="w")
        
        carpeta_origen_entry = ttk.Entry(self.configuracion_tab, textvariable=self.carpeta_archivos, width=40)
        carpeta_origen_entry.grid(row=9, column=2, columnspan=2, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="Carpeta Origen Imagen:").grid(row=10, column=1, pady=5, sticky="w", padx= espaciado_lateral)

        # Mostrar la imagen "folder.png" al lado del campo "Carpeta Origen Imagen"
        folder_image = Image.open("Icons/folder.png")
        folder_image = folder_image.resize((20, 20))  # Redimensiona la imagen
        folder_icon = ImageTk.PhotoImage(folder_image)
        
        # Asignar la función seleccionar_carpeta_imagen al evento de clic del botón
        def seleccionar_carpeta_imagen():
            selected_folder = filedialog.askdirectory()
            if selected_folder:
                carpeta_origen_imagen_entry.delete(0, tk.END)  # Borrar el contenido actual del Entry
                carpeta_origen_imagen_entry.insert(0, selected_folder)  # Insertar la nueva ruta seleccionada

        folder_label_imagen = ttk.Button(self.configuracion_tab, image=folder_icon, command=seleccionar_carpeta_imagen)
        folder_label_imagen.image = folder_icon
        folder_label_imagen.grid(row=10, column=4, padx=(5, 0), pady=5, sticky="w")
        
        carpeta_origen_imagen_entry = ttk.Entry(self.configuracion_tab, width=40, textvariable=self.carpeta_imagenes)
        carpeta_origen_imagen_entry.grid(row=10, column=2, columnspan=2, pady=5, sticky="w")

        
        save_image = customtkinter.CTkImage(Image.open("Icons/save.png").resize((100,100), Image.Resampling.LANCZOS))
        boton_save = customtkinter.CTkButton(self.configuracion_tab, text="Guardar Configuración", border_color="#AFACAC", border_width=1,   corner_radius=5,font=("Helvetica", 14), text_color="#000000", fg_color="#FFFFFF", hover_color="#D9F3FF", width=120, height=20, compound="left", image= save_image, command=self.guardar_configuracion)
        boton_save.grid(row=11, column=2, padx=(10,30), pady=10)
        
        cambiar_contraseña_button = ttk.Button(self.configuracion_tab, text="Cambiar Contraseña", command=self.abrir_ventana_cambio_contraseña)
        cambiar_contraseña_button.grid(row=12, rowspan=3, columnspan=2, padx=10, pady=(10,5), sticky="s")

#METODO PARA ENVIAR EL JSON DE DATOS Y ENVIARLO AL WEBSERVICE
#Metodo para envío de datos a WebService
    def enviar_data(self, data, url, archivo, es_imagen=False):
        max_connection_attempts = 3  # Número máximo de intentos de conexión
        max_request_attempts = 3  # Número máximo de intentos de solicitud

        connection_attempts = 0
        while connection_attempts < max_connection_attempts:
            try:
                # Intentar conexión
                token_response = requests.post(
                    self.token_url.get(),
                    auth=HTTPBasicAuth(self.client_id.get(), self.client_secret.get()),
                    data={
                        "grant_type": "password",
                        "username": self.username.get(),
                        "password": self.password.get()
                    }
                )

                if token_response.status_code == 200:
                    access_token = token_response.json().get("access_token")
                    if access_token:
                        headers = {
                            "Authorization": f"Bearer {access_token}",
                            "Content-Type": "application/json"
                        }

                        request_attempts = 0
                        while request_attempts < max_request_attempts:
                            # Intentar enviar la solicitud de datos
                            self.response = requests.post(url, json=data, headers=headers)
                            if self.response.status_code == 200:
                                try:
                                    self.error = False
                                    self.json_response = self.response.json()
                                    self.envio_exitoso += 1
                                    self.tree.tag_configure('verde', background='lightgreen')
                                    self.tree.item(self.item_id, tags=('verde',))
                                    return  # Salir del método si la solicitud es exitosa
                                except requests.exceptions.JSONDecodeError:
                                    try:
                                        self.error=False
                                        # Intenta analizar la respuesta como XML
                                        xml_response = ET.fromstring(self.response.text)
                                        return  # Salir del método si la solicitud es exitosa
                                    except ET.ParseError:
                                        messagebox.showerror("La respuesta no es ni JSON ni XML válido. Contenido de la respuesta:", response.text)
                            time.sleep(1)  # Esperar 1 segundo antes de reintentar
                            request_attempts += 1

                        # Si se agotaron los intentos de solicitud, mover el archivo a la carpeta de errores
                        if es_imagen:
                            self.mover_a_carpeta_errores(archivo, es_imagen=True)
                        else:
                            self.mover_a_carpeta_errores(archivo, es_imagen=False)
                        self.envio_fallido += 1
                        self.tree.tag_configure('rojo', background='#FA5656')
                        self.tree.item(self.item_id, tags=('rojo',))
                        self.update_contadores()
                        return  # Salir del bucle de conexión si la solicitud falló
                    else:
                        # Si no se pudo obtener el token de acceso, mover el archivo a la carpeta de errores
                        if es_imagen:
                            self.mover_a_carpeta_errores(archivo, es_imagen=True)
                        else:
                            self.mover_a_carpeta_errores(archivo, es_imagen=False)
                        self.envio_fallido += 1
                        self.tree.tag_configure('rojo', background='#FA5656')
                        self.tree.item(self.item_id, tags=('rojo',))
                        self.update_contadores()
                        break  # Salir del bucle de conexión si la solicitud falló
                else:
                    # Si no se pudo obtener el token de acceso, mover el archivo a la carpeta de errores
                    if es_imagen:
                        self.mover_a_carpeta_errores(archivo, es_imagen=True)
                    else:
                        self.mover_a_carpeta_errores(archivo, es_imagen=False)
                    self.envio_fallido += 1
                    self.tree.tag_configure('rojo', background='#FA5656')
                    self.tree.item(self.item_id, tags=('rojo',))
                    self.update_contadores()
                    break  # Salir del bucle de conexión si la solicitud falló
            except ConnectionError:
                # Si falla la conexión, intentar nuevamente después de un breve retraso
                time.sleep(1)
                connection_attempts += 1

        # Si se agotaron los intentos de conexión, mostrar un mensaje de error y mover el archivo a la carpeta de errores
        self.tree.tag_configure('rojo', background='#FA5656')
        self.tree.item(self.item_id, tags=('rojo',))
        self.update_contadores()
        self.error = True
        messagebox.showerror("Error de conexión", "No se pudo establecer conexión con el servidor.")
        if es_imagen:
            self.mover_a_carpeta_errores(archivo, es_imagen=True)
        else:
            self.mover_a_carpeta_errores(archivo, es_imagen=False)

    #Metodo para verificar si hay internet
    def verificar_conexion(self):
        try:
            # Intentar hacer una solicitud a un sitio web conocido
            requests.get("http://www.google.com", timeout=1)
            return True
        except requests.ConnectionError:
            return False

#Procesamiento de archivos e imagenes
    #Metodo para mover errores a carpeta predeterminada
    def mover_a_carpeta_errores(self, archivo, es_imagen=False):
        carpeta_errores = self.carpeta_procesados_data_e if not es_imagen else self.carpeta_procesados_img_e
        carpeta_errores = os.path.join(carpeta_errores)
        if not os.path.exists(carpeta_errores):
            os.makedirs(carpeta_errores)
        archivo_con_error = os.path.join(carpeta_errores, os.path.basename(archivo))
        if os.path.exists(archivo_con_error):
            os.remove(archivo_con_error)
        try:
            os.rename(archivo, archivo_con_error)
        except: pass

    #Metodo para procesar archivo TXT
    def procesar_archivo(self, archivo):

        if not self.verificar_conexion():
            messagebox.showerror("Error de conexión", "No hay conexión a Internet.")
            return

        try:
            carpeta_procesados_data = os.path.join(self.carpeta_procesados_data)
            if not os.path.exists(carpeta_procesados_data):
                os.makedirs(carpeta_procesados_data)

            nuevo_nombre = os.path.join(carpeta_procesados_data, os.path.basename(archivo))
            # Verificar si el archivo de destino ya existe y eliminarlo
            if os.path.exists(nuevo_nombre):
                os.remove(nuevo_nombre)
            
            with open(archivo, "r") as f:
                line = f.readline().strip()
                datos = line.split("|")

                if len(datos) != 10:
                    raise ValueError("La estructura del archivo no es válida")

                SKU, Packtype, Tipodepaquete, Cantidad, CantidadInner, Largo, Ancho, Alto, Peso, Descripcion = datos

                # Convertir los valores a números (Largo, Ancho, Alto y Peso)
                Largo = float(Largo)
                Ancho = float(Ancho)
                Alto = float(Alto)
                Peso = float(Peso)
                Cantidad = int(Cantidad)
                CantidadInner = int(CantidadInner)
                fecha = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
                if Cantidad == 0:
                    Cantidad=1
                else: Cantidad=Cantidad
                
                packkey = f"{SKU}_{Cantidad}"

                if CantidadInner == 0:
                    CantidadInner=1
                else: CantidadInner=CantidadInner
                
                if CantidadInner>Cantidad:
                    raise ValueError(f'La cantidad de inner ({CantidadInner}) del articulo {SKU} es mayor a la cantidad de cajas ({Cantidad})')

                if Packtype == "Unidad-UOM3":
                    data = {
                        "packkey": packkey,
                        "packdescr": Descripcion,
                        "packuom1": "CJ",
                        "packuom3": Tipodepaquete,
                        "casecnt": Cantidad,
                        "qty": 1,
                        "widthuom3": Ancho,
                        "lengthuom3": Largo,
                        "heightuom3": Alto,
                        "weightuom3": Peso,
                        "widthuom1": Ancho,
                        "lengthuom1": Largo,
                        "heightuom1": Alto,
                        "weightuom1": Peso,
                        "pallethi": 1,
                        "palletti": 1,
                        "ext_udf_str1": SKU,
                        "ext_udf_str2": Tipodepaquete
                    }
                    #print(data)
                    f.close()
                    self.item_id=self.tree.insert('', 'end', values=(packkey, Packtype, Tipodepaquete, Cantidad , CantidadInner,Largo, Ancho, Alto, Peso, fecha))
                    self.enviar_data(data, self.api_url.get(), archivo, es_imagen=False)
                    self.actualizar_log(packkey, es_imagen=False)
                    self.exportar_excel(packkey, Packtype, Tipodepaquete, Cantidad, CantidadInner, Largo, Ancho, Alto, Peso, fecha)
                elif Packtype == "Caja-UOM1" or Packtype == "Caja2-UOM1" or Packtype == "Caja3-UOM1" or Packtype == "Caja4-UOM1" or Packtype == "Caja5-UOM1" :
                    data = {
                        "packkey": packkey,
                        "packdescr": Descripcion,
                        "packuom1": Tipodepaquete,
                        "casecnt": Cantidad,
                        "widthuom1": Ancho,
                        "lengthuom1": Largo,
                        "heightuom1": Alto,
                        "weightuom1": Peso,
                        "pallethi": 1,
                        "palletti": 1,
                        "ext_udf_str1": SKU,
                        "ext_udf_str2": Tipodepaquete
                    }
                    #print(data)
                    f.close()
                    self.item_id=self.tree.insert('', 'end', values=(packkey, Packtype, Tipodepaquete, Cantidad , CantidadInner, Largo, Ancho, Alto, Peso, fecha))
                    self.enviar_data(data, self.api_url.get(), archivo, es_imagen=False)
                    self.actualizar_log(packkey, es_imagen=False)
                    self.exportar_excel(packkey, Packtype, Tipodepaquete, Cantidad,CantidadInner, Largo, Ancho, Alto, Peso, fecha)
                elif Packtype == "Subcaja-UOM2":
                    data = {
                        "packkey": packkey,
                        "packdescr": Descripcion,
                        "packuom2": Tipodepaquete,
                        "innerpack": CantidadInner,
                        "widthuom2": Ancho,
                        "lengthuom2": Largo,
                        "heightuom2": Alto,
                        "weightuom2": Peso,
                        "pallethi": 1,
                        "palletti": 1,
                        "ext_udf_str1": SKU,
                        "ext_udf_str2": Tipodepaquete
                    }
                    #print(data)
                    f.close()
                    self.item_id=self.tree.insert('', 'end', values=(packkey, Packtype, Tipodepaquete, Cantidad , CantidadInner,Largo, Ancho, Alto, Peso, fecha))
                    self.enviar_data(data, self.api_url.get(), archivo, es_imagen=False)
                    self.actualizar_log(packkey, es_imagen=False)
                    self.exportar_excel(packkey, Packtype, Tipodepaquete, Cantidad, CantidadInner, Largo, Ancho, Alto, Peso, fecha)
            
                if not self.error:
                    carpeta_procesados_data = os.path.join(self.carpeta_procesados_data)
                    if not os.path.exists(carpeta_procesados_data):
                        os.makedirs(carpeta_procesados_data)
                    
                    # Mueve el archivo procesado a la carpeta "procesados"
                    nuevo_nombre = os.path.join(self.carpeta_procesados_data, os.path.basename(archivo))
                    # Cerrar el archivo antes de intentar moverlo
                    os.rename(archivo, nuevo_nombre)
                self.update_contadores()
                #print(f"Archivo procesado: {archivo}")


        except ValueError as ve:
            # Manejar el error si la estructura del archivo no es válida
            messagebox.showerror("Error", f"Error al procesar el archivo {archivo}: {str(ve)}")
            
            carpeta_procesados_data_e = os.path.join(self.carpeta_procesados_data_e)
            if not os.path.exists(carpeta_procesados_data_e):
                os.makedirs(carpeta_procesados_data_e)

            # Mueve el archivo con error a la carpeta de errores
            nuevo_nombre_error = os.path.join(carpeta_procesados_data_e, os.path.basename(archivo))
            f.close()
            
            if os.path.exists(nuevo_nombre_error):
                os.remove(nuevo_nombre_error)
            
            os.rename(archivo, nuevo_nombre_error)
    
    #Metodo para procesar archivo JPG
    def procesar_imagen(self, ruta_imagen):
        try:
            
            carpeta_procesados_img = os.path.join(self.carpeta_procesados_img)
            if not os.path.exists(carpeta_procesados_img):
                os.makedirs(carpeta_procesados_img)

            nuevo_nombre = os.path.join(carpeta_procesados_img, os.path.basename(ruta_imagen))

            # Verificar si el archivo de destino ya existe y eliminarlo
            if os.path.exists(nuevo_nombre):
                os.remove(nuevo_nombre)

            with open(ruta_imagen, "rb") as img_file:
                # Leer la imagen en bytes
                img_bytes = img_file.read()
                SKU= os.path.basename(ruta_imagen).split('_')[0]
                Propietario = os.path.basename(ruta_imagen).split('_')[2]
                
                if Propietario=="Mavesa":
                    cod_propietario="0001"
                elif Propietario=="Internaconsa":
                    cod_propietario="0005"
                else:
                    cod_propietario="0002"
                    
                # Codificar la imagen en base64
                img_base64 = base64.b64encode(img_bytes).decode('utf-8')
                # Construir el JSON de la imagen
                json_imagen = {
                    "item": {
                        "attrs": {
                            "attr": [
                                {"name": "storer", "value": cod_propietario},
                                {"name": "sku", "value": os.path.basename(ruta_imagen).split('_')[0]},  # Obtener el "CODIGO" del nombre de la imagen
                                {"name": "uom", "value": os.path.basename(ruta_imagen).split('_')[1].split('.')[0]}  # Obtener el "UN" del nombre de la imagen
                            ]
                        },
                        "resrs": {
                            "res": [
                                {
                                    "filename": os.path.basename(ruta_imagen),
                                    "base64": img_base64
                                }
                            ]
                        },
                        "acl": {
                            "name": "Public"
                        },
                        "entityName": "SCE_Product_Image"
                    }
                }
                # Enviar el JSON al servicio de imágenes
                #if not self.error:
                img_file.close()
                self.enviar_data(json_imagen, self.url_api_image.get(), ruta_imagen, es_imagen=True)
                self.actualizar_log(SKU, es_imagen=True)
                
                carpeta_procesados_img = os.path.join(self.carpeta_procesados_img)
                if not os.path.exists(carpeta_procesados_img):
                    os.makedirs(carpeta_procesados_img)
                
                nuevo_nombre = os.path.join(carpeta_procesados_img, os.path.basename(ruta_imagen))
                # Cerrar el archivo antes de intentar moverlo
                try: 
                    os.rename(ruta_imagen, nuevo_nombre)
                except: pass
                    #print("Ya se ha movido el archivo")
                #print(f"Imagen procesada: {ruta_imagen}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar la imagen:", f"Error: {str(e)}")

    #Metodo para que siempre tome el archivo mas antiguo de la lista
    def obtener_archivo_mas_antiguo(self, carpeta, extension=None, es_imagen=False):
        time.sleep(0.5)
        archivos = [f for f in os.listdir(carpeta) if f.endswith(extension)] if extension else os.listdir(carpeta)
        if not archivos:
            return None
        try:
            for archivo in archivos:

                if extension == ".jpg" and es_imagen:
                    datetime.strptime(archivo.split("_")[3].replace('.jpg', ''), "%Y%m%d%H%M%S")
                elif extension == ".txt" and not es_imagen:
                    datetime.strptime(archivo.split("_")[1].replace('.txt', ''), "%Y%m%d%H%M%S")
        except Exception as e:
            # Manejar el error individualmente para cada archivo
            messagebox.showerror("Error", f"El archivo ({archivo}): no cumple con la estructura definida {str(e)}")
            #print(datetime.strptime(archivo.split("_")[3].replace('.jpg', ''), "%Y%m%d%H%M%S"))

            carpeta_errores = self.carpeta_procesados_data_e if not es_imagen else self.carpeta_procesados_img_e
            carpeta_errores = os.path.join(carpeta_errores)

            if not os.path.exists(carpeta_errores):
                os.makedirs(carpeta_errores)

            # Mueve el archivo con error a la carpeta de errores
            archivo_con_error = os.path.join(carpeta_errores, archivo)
            
            if os.path.exists(archivo_con_error):
                os.remove(archivo_con_error)

            try:
                os.rename(os.path.join(carpeta, archivo), archivo_con_error)
            except: pass   
            time.sleep(0.5)

            return None

        # Ordenar archivos después de asegurarse de que todos cumplen con el formato
        archivos.sort(key=lambda x: datetime.strptime(x.split("_")[3].replace('.jpg', ''), "%Y%m%d%H%M%S")) if extension == ".jpg" and es_imagen else archivos.sort(key=lambda x: datetime.strptime(x.split("_")[1].replace('.txt', ''), "%Y%m%d%H%M%S"))
        
        if archivos:
            # Si no hubo error, devolver el archivo más antiguo
            return os.path.join(carpeta, archivos[0])
        else:
            return None

    #Metodo para que contantemente esté tomando archivos si los hay
    def procesar_archivos_continuamente(self):
        while self.ejecutar:
            try: 
                archivo_txt = self.obtener_archivo_mas_antiguo(self.carpeta_archivos.get(), ".txt", es_imagen=False)
                archivo_img = self.obtener_archivo_mas_antiguo(self.carpeta_imagenes.get(), ".jpg", es_imagen=True)  # Ajustar la extensión
            

                if archivo_txt:
                    self.error=True
                    self.procesar_archivo(archivo_txt)
                elif archivo_img:
                    self.error=True
                    self.procesar_imagen(archivo_img)
            except: pass

#Actualización de interfaz
    #Actualizar contadores de envios exitosos y erroneos
    def update_contadores(self):
        self.label_envio_exitoso.config(text=f"Envíos exitosos: {self.envio_exitoso}")
        self.label_envio_fallido.config(text=f"Envíos fallidos: {self.envio_fallido}")

    #Actualizar el campo de respuesta
    def actualizar_log(self, SKU, es_imagen=False):
        fecha = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        self.response_entry.tag_config('warning', foreground="#FA5656")
        self.response_entry.tag_config('ok', foreground="green")
        self.response_entry.tag_config('image', foreground="blue")

        if self.error and es_imagen:
            self.response_entry.config(state=tk.NORMAL)  # Habilita la edición temporalmente
            self.response_entry.insert(tk.END, f"{fecha}  SKU={SKU}, Respuesta WS:  {self.response.text}\n", 'warning')
            self.response_entry.config(state=tk.DISABLED)  # Deshabilita la edición temporalmente 
        elif self.error and not es_imagen:
            self.response_entry.config(state=tk.NORMAL)  # Habilita la edición temporalmente
            self.response_entry.insert(tk.END, f"{fecha}  SKU={SKU}, Respuesta WS:  {self.response.text}\n", 'warning')
            self.response_entry.config(state=tk.DISABLED)  # Deshabilita la edición temporalmente 
        elif not self.error and not es_imagen:
            self.response_entry.config(state=tk.NORMAL)  # Habilita la edición temporalmente
            self.response_entry.insert(tk.END, f"{fecha}  packkey={SKU}, Respuesta WS:  Los datos fueron enviados exitosamente\n", 'ok')
            self.response_entry.config(state=tk.DISABLED)  # Deshabilita la edición temporalmente 
        else:
            self.response_entry.config(state=tk.NORMAL)  # Habilita la edición temporalmente
            self.response_entry.insert(tk.END, f"{fecha}  packkey={SKU}, Respuesta WS:  La imagen fue enviada exitosamente\n", 'image')
            self.response_entry.config(state=tk.DISABLED)  # Deshabilita la edición temporalmente 
        self.response_entry.see(tk.END)  # Desplaza la vista al final del texto
        self.tree.yview_moveto(1.0)  # Desplaza la vista hacia el final de la tabla

    #Exportar el excel con los datos medidos
    def exportar_excel(self, SKU, Packtype, Tipodepaquete, Cantidad, CantidadInner, Largo, Ancho, Alto, Peso, fecha):
        self.ruta_exportacion = ""
        fecha_actual = datetime.now().strftime("%d-%m-%Y")

        # Configurar la carpeta de destino predeterminada
        default_folder = "Export"

        # Usar la carpeta predeterminada si self.ruta_exportacion está vacía
        if not self.ruta_exportacion or not Path(self.ruta_exportacion).exists() or not Path(self.ruta_exportacion).is_dir():
            self.ruta_destino = Path(default_folder)
            if not self.ruta_destino.exists():
                os.makedirs(self.ruta_destino)
        else:
            self.ruta_destino = Path(self.ruta_exportacion)

        # Corregir el nombre de la variable
        nombre_archivo = f"CubiScan_{fecha_actual}.xlsx"
        ruta_completa = self.ruta_destino / nombre_archivo

        # Verificar si el archivo ya existe
        if ruta_completa.exists():
            workbook = openpyxl.load_workbook(ruta_completa)
            worksheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Medidas"
            # Encabezados
            encabezados = ["SKU", "Packtype", "Tipodepaquete", "Cajas", "Inner","Largo", "Ancho", "Alto", "Peso", "Fecha"]
            for col_num, encabezado in enumerate(encabezados, 1):
                worksheet.cell(row=1, column=col_num, value=encabezado)

        # Agregar nueva fila
        nueva_fila = [SKU, Packtype, Tipodepaquete, Cantidad, CantidadInner,  Largo, Ancho, Alto, Peso, fecha]
        worksheet.append(nueva_fila)

        # Guardar el archivo Excel
        workbook.save(ruta_completa)

    #Exportar lo almacenado en el campo de respuestas
    def exportar_log(self):
        # Obtener la fecha actual en formato %d-%m-%Y
        fecha_actual = datetime.now().strftime("%d-%m-%Y")
        
        # Nombre del archivo con la fecha actual
        file_name = f"log_{fecha_actual}.txt"
        file_path = os.path.join("Log", file_name)

        # Obtener el texto actual en el Entry
        text_to_export = self.response_entry.get("1.0", "end-1c")

        # Verificar si la carpeta "Log" existe, y si no, crearla
        log_folder = "Log"
        if not os.path.exists(log_folder):
            os.makedirs(log_folder)

        
        # Si el archivo ya existe, agregar nuevo contenido
        if os.path.exists(file_path):
            with open(file_path, "a") as file:
                # Agregar nueva línea y el texto actual
                file.write(text_to_export)
        else:
            # Si el archivo no existe, crear uno nuevo
            with open(file_path, "w") as file:
                file.write(text_to_export)

#METODOS PARA INICIAR Y DETENER EL PROCESAMIENTO DE DATOS
    def iniciar_proceso(self):
        # Inicia un hilo para ejecutar el procesamiento en segundo plano
        self.boton_iniciar.configure(state="disabled", fg_color="#2DE524")
        self.boton_detener.configure(state="normal", fg_color="#FFFFFF")
        self.ejecutar=True
        Thread(target=self.procesar_archivos_continuamente).start()

    def detener_proceso(self):
        try:
            self.boton_iniciar.configure(state="normal", fg_color="#FFFFFF")
            self.boton_detener.configure(state="disabled", fg_color="#FC0909")
            # Detiene el hilo de procesamiento
            self.ejecutar = False
            #messagebox.showinfo("Proceso detenido", "El proceso ha sido detenido exitosamente.")
        except Exception as e:
            messagebox.showerror("Error al detener el proceso", f"Error: {str(e)}")

    def ejecutar_interfaz(self):
        # Ejecutar la interfaz gráfica
        root.mainloop()



if __name__ == "__main__":
    root = tk.Tk()
    root.resizable(False,False)
    app = ProcesadorArchivos(root)
    app.ejecutar_interfaz()
